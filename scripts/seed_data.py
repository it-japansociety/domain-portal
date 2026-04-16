"""
seed_data.py
Handles .csv or .xlsx automatically.
Extracts domain from email addresses (user@contoso.com -> contoso.com).
"""

import argparse
import csv
import re
import sys
from pathlib import Path

try:
    from azure.data.tables import TableClient, UpdateMode
except ImportError:
    sys.exit("Missing dependency: run  pip install azure-data-tables")


def extract_domain(raw: str) -> str:
    """Handles plain domains or full email addresses."""
    raw = (raw or "").strip().lower()
    raw = re.sub(r"^https?://", "", raw)
    raw = re.sub(r"^www\.", "", raw)
    if "@" in raw:
        raw = raw.split("@")[1]
    return raw.split("/")[0].split("?")[0].split(":")[0]


def domain_to_row_key(domain: str) -> str:
    return domain.replace(".", "|")


def find_column(headers: list, *candidates) -> str | None:
    lower = [h.lower().strip() for h in headers]
    for name in candidates:
        if name.lower() in lower:
            return headers[lower.index(name.lower())]
    return None


def read_file(path: Path) -> tuple:
    """Returns (headers, rows) for either .csv or .xlsx"""
    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            sys.exit("Excel file detected — run: pip install openpyxl")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        all_rows = [[str(c if c is not None else "") for c in row]
                    for row in ws.iter_rows(values_only=True)]
        wb.close()
        if not all_rows:
            sys.exit("File is empty.")
        return all_rows[0], all_rows[1:]

    else:  # treat as CSV
        with open(path, newline="", encoding="utf-8-sig") as f:
            all_rows = list(csv.reader(f))
        if not all_rows:
            sys.exit("File is empty.")
        return all_rows[0], all_rows[1:]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, help="Path to .csv or .xlsx file")
    parser.add_argument("--connection-string", required=True)
    parser.add_argument("--table", default="domaincodes")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        sys.exit(f"File not found: {path}")

    print(f"\n📂  Loading: {path.name}")
    headers, rows = read_file(path)
    print(f"    Columns found: {headers}")

    # Find columns — accepts email OR domain column names
    col_email = find_column(headers,
        "email", "emailaddress", "email address",
        "domain", "url"
    )
    col_code = find_column(headers,
        "code", "codes", "access_code", "accesscode", "new code", "newcode"
    )
    col_label = find_column(headers,
        "label", "organization", "org", "name", "company"
    )

    if not col_email:
        sys.exit(f"❌  No email or domain column found.\n    Headers: {headers}")
    if not col_code:
        sys.exit(f"❌  No code column found.\n    Headers: {headers}")

    idx_email = headers.index(col_email)
    idx_code  = headers.index(col_code)
    idx_label = headers.index(col_label) if col_label else None

    print(f"    Email/Domain col : '{col_email}'")
    print(f"    Code col         : '{col_code}'")
    print(f"    Label col        : '{col_label or 'not found'}'")

    records = []
    skipped = []

    for i, row in enumerate(rows, start=2):
        # Pad short rows
        row = list(row) + [""] * (len(headers) - len(row))

        raw_email = row[idx_email].strip()
        raw_code  = row[idx_code].strip()
        raw_label = row[idx_label].strip() if idx_label is not None else ""

        if not raw_email or not raw_code:
            skipped.append((i, raw_email, "Missing email or code"))
            continue

        domain = extract_domain(raw_email)

        if not domain or "." not in domain:
            skipped.append((i, raw_email, f"Invalid domain extracted: '{domain}'"))
            continue

        records.append({
            "PartitionKey": "domains",
            "RowKey":       domain_to_row_key(domain),
            "domain":       domain,
            "code":         raw_code,
            "label":        raw_label,
        })

    print(f"\n✅  Valid records : {len(records)}")
    if skipped:
        print(f"⚠️   Skipped       : {len(skipped)}")
        for row_num, val, reason in skipped[:10]:
            print(f"     Row {row_num}: '{val}' — {reason}")

    print("\nSample (first 5):")
    for r in records[:5]:
        print(f"  {r['domain']:<40}  →  {r['code']}")

    if args.dry_run:
        print("\n🔍  Dry run complete — no data written.")
        return

    confirm = input(f"\nUpload {len(records)} records to table '{args.table}'? [y/N] ").strip().lower()
    if confirm != "y":
        print("Aborted.")
        return

    client = TableClient.from_connection_string(args.connection_string, args.table)
    try:
        client.create_table()
        print(f"Created table: {args.table}")
    except Exception:
        print(f"Table already exists: {args.table}")

    success = 0
    errors  = 0
    for rec in records:
        try:
            client.upsert_entity(rec, mode=UpdateMode.REPLACE)
            success += 1
        except Exception as e:
            print(f"  ❌  Failed '{rec['domain']}': {e}")
            errors += 1

    print(f"\n🏁  Done — {success} upserted, {errors} errors.")


if __name__ == "__main__":
    main()